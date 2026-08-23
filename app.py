#!/usr/bin/env python3
"""
app.py — Redirector Flask Application Entrypoint

Hardened Application Architecture:
  - Session Auth & Protection: HTTP-only session tokens & uniform endpoint authentication
  - SSRF Defense: Hostname & IP resolution filtering (blocking private IPs & cloud metadata)
  - XML Security: defusedxml entity-expansion mitigation for sitemaps
  - Payload Protection: 16 MB MAX_CONTENT_LENGTH capping
  - Token Encryption: Fernet AES key encryption at rest for Shopify API tokens
  - Database & Queue: SQLite database engine with persistent background task queueing
"""

import os
import re
import io
import csv
import json
import time
import secrets
import threading
import urllib.request
import urllib.error
from urllib.parse import urlparse, unquote
from concurrent.futures import ThreadPoolExecutor

from flask import Flask, request, jsonify, Response, abort, send_file, render_template, send_from_directory, session

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import defusedxml.ElementTree as SafeET
except ImportError:
    import xml.etree.ElementTree as SafeET

from core import auth, matcher, live_checker, catalog_fetcher, db, validators
from core.logger import logger

# App Secret Key for HTTP-only session cookies & multi-worker WSGI persistence
APP_SECRET = os.environ.get("APP_SECRET") or os.environ.get("SECRET_KEY") or secrets.token_hex(24)

# ---------------------------------------------------------------------------
# App & Database Initialization
# ---------------------------------------------------------------------------
CFG = auth.load_config()
SHOP = CFG["shop"]
API_VERSION = "2024-10"
CONFIGURED = auth.is_configured(CFG)

# Initialize SQLite Database & Auto-Migrate legacy JSON files
db.init_db()
logger.info("SQLite Database initialized at data/redirector.db")

DATA_LOCK = threading.Lock()


def get_products():
    return db.get_all_products()


def get_patterns():
    return db.get_all_patterns()


# ---------------------------------------------------------------------------
# Shopify Admin API Helpers & Health Monitor
# ---------------------------------------------------------------------------
SHOPIFY_API_HEALTH = {
    "status": "healthy",  # "healthy", "throttled", "rate_limited", "error"
    "call_limit": "0/40",
    "used": 0,
    "max": 40,
    "last_code": 200,
    "last_updated": 0,
    "message": "Shopify API Healthy",
}


def update_shopify_health(status_code, headers=None, err_msg=None):
    now = time.time()
    SHOPIFY_API_HEALTH["last_updated"] = now
    SHOPIFY_API_HEALTH["last_code"] = status_code

    if headers:
        limit_hdr = headers.get("X-Shopify-Shop-Api-Call-Limit") or headers.get("x-shopify-shop-api-call-limit")
        if limit_hdr and "/" in str(limit_hdr):
            SHOPIFY_API_HEALTH["call_limit"] = str(limit_hdr)
            try:
                parts = str(limit_hdr).split("/")
                used = int(parts[0])
                max_cap = int(parts[1])
                SHOPIFY_API_HEALTH["used"] = used
                SHOPIFY_API_HEALTH["max"] = max_cap
                ratio = used / max_cap if max_cap > 0 else 0

                if ratio >= 0.85:
                    SHOPIFY_API_HEALTH["status"] = "throttled"
                    SHOPIFY_API_HEALTH["message"] = f"Shopify API Throttled ({limit_hdr} calls used)"
                else:
                    SHOPIFY_API_HEALTH["status"] = "healthy"
                    SHOPIFY_API_HEALTH["message"] = f"Shopify API Healthy ({limit_hdr})"
            except Exception:
                pass

    if status_code in (429, 409):
        SHOPIFY_API_HEALTH["status"] = "rate_limited"
        SHOPIFY_API_HEALTH["message"] = f"Shopify API Rate Limited ({status_code}) — Auto Backoff Active"
    elif status_code in (401, 403):
        SHOPIFY_API_HEALTH["status"] = "error"
        SHOPIFY_API_HEALTH["message"] = f"Shopify Auth Error ({status_code})"
    elif status_code >= 500:
        SHOPIFY_API_HEALTH["status"] = "error"
        SHOPIFY_API_HEALTH["message"] = f"Shopify Server Issue ({status_code})"


def shopify_request(method, endpoint, body=None, _retry=True, _attempts=0):
    url = f"https://{SHOP}/admin/api/{API_VERSION}/{endpoint}"
    data = json.dumps(body).encode("utf-8") if body is not None else None
    try:
        token = auth.get_valid_token(CFG)
    except RuntimeError as e:
        logger.error(f"Failed to retrieve valid Shopify access token: {e}")
        update_shopify_health(0, err_msg=str(e))
        return 0, {"error": str(e)}

    req = urllib.request.Request(
        url, data=data, method=method,
        headers={"Content-Type": "application/json", "X-Shopify-Access-Token": token},
    )
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read().decode("utf-8")
            update_shopify_health(resp.status, headers=resp.headers)
            return resp.status, (json.loads(raw) if raw else {})
    except urllib.error.HTTPError as e:
        update_shopify_health(e.code, headers=e.headers, err_msg=str(e))
        if e.code == 401 and _retry:
            auth.get_valid_token(CFG, force_refresh=True)
            return shopify_request(method, endpoint, body, _retry=False, _attempts=_attempts)
        elif e.code in (429, 409) and _attempts < 4:
            retry_after = e.headers.get("Retry-After")
            wait_time = float(retry_after) + 0.5 if (retry_after and retry_after.replace('.', '', 1).isdigit()) else (1.5 * (_attempts + 1))
            logger.warning(f"Shopify Admin API {e.code} Rate Limited. Backing off for {wait_time:.1f}s (attempt {_attempts + 1}/4)...")
            time.sleep(wait_time)
            return shopify_request(method, endpoint, body, _retry=_retry, _attempts=_attempts + 1)
        raw = e.read().decode("utf-8")
        return e.code, (json.loads(raw) if raw else {})
    except urllib.error.URLError as e:
        logger.error(f"URLError contacting Shopify API endpoint {endpoint}: {e.reason}")
        update_shopify_health(0, err_msg=str(e.reason))
        return 0, {"error": str(e.reason)}


def to_path(url_or_path):
    if url_or_path.startswith("http"):
        return urlparse(url_or_path).path
    return url_or_path if url_or_path.startswith("/") else "/" + url_or_path


def create_or_update_redirect(from_url, to_url, overwrite_existing=False):
    path = to_path(from_url)
    target = to_path(to_url)

    status, resp = shopify_request("POST", "redirects.json", {"redirect": {"path": path, "target": target}})
    if status in (200, 201):
        logger.info(f"Successfully created Shopify redirect #{resp['redirect']['id']}: {path} -> {target}")
        return True, resp["redirect"]["id"], None, "applied"

    status2, existing = shopify_request("GET", f"redirects.json?path={path}")
    if status2 == 200 and existing.get("redirects"):
        existing_red = existing["redirects"][0]
        rid = existing_red["id"]
        existing_target = existing_red.get("target", "")

        if not overwrite_existing:
            logger.info(f"Skipped existing redirect on Shopify #{rid}: {path} -> {existing_target}")
            return False, rid, f"Skipped: redirect already exists on Shopify (redirects to {existing_target})", "skipped"

        status3, resp3 = shopify_request(
            "PUT", f"redirects/{rid}.json",
            {"redirect": {"id": rid, "path": path, "target": target}},
        )
        if status3 == 200:
            logger.info(f"Updated existing Shopify redirect #{rid}: {path} -> {target}")
            return True, rid, None, "applied"
        return False, None, resp3, "error"

    logger.warning(f"Failed to create Shopify redirect ({status}): {resp}")
    return False, None, resp, "error"


# ---------------------------------------------------------------------------
# Defused XML Sitemap Parser & Background Scanner
# ---------------------------------------------------------------------------
def fetch_sitemap_urls(sitemap_url, timeout=10):
    if not validators.is_safe_public_url(sitemap_url, shop_domain=SHOP):
        raise validators.ValidationError("Access to private, loopback, or internal IP address is blocked (SSRF Protection).")

    req = urllib.request.Request(sitemap_url, headers={"User-Agent": "Mozilla/5.0 (compatible; BrokenLinkChecker/1.0)"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()

    # Defused XML parsing (mitigates entity expansion & DTD bombs)
    root = SafeET.fromstring(raw)
    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    locs = [el.text.strip() for el in root.findall(".//sm:loc", ns) if el.text]

    if root.tag.endswith("sitemapindex"):
        all_urls = []
        for child_sitemap in locs[:20]:
            try:
                all_urls.extend(fetch_sitemap_urls(child_sitemap, timeout=timeout))
            except Exception as e:
                logger.warning(f"Failed to parse child sitemap {child_sitemap}: {e}")
                continue
        return all_urls
    return locs


def add_new_broken_url(broken_url, products=None, patterns=None, inverted_index=None, existing_urls_set=None):
    if existing_urls_set is not None:
        if broken_url in existing_urls_set:
            return None
    else:
        all_matches = db.get_all_matches()
        if any(m["broken_url"] == broken_url for m in all_matches):
            return None

    prods = products if products is not None else get_products()
    pats = patterns if patterns is not None else get_patterns()
    inv_idx = inverted_index if inverted_index is not None else matcher.build_product_index(prods)

    matches = matcher.resolve_matches_for_url(broken_url, prods, pats, inverted_index=inv_idx)
    new_id = db.get_next_match_id()
    row = {
        "id": new_id,
        "broken_url": broken_url,
        "extracted_name": matcher.extract_candidate_name(broken_url),
        "matches": matches,
        "status": "pending",
        "chosen_index": 0,
        "source": "watchlist_scan",
    }
    db.upsert_match(row)
    if existing_urls_set is not None:
        existing_urls_set.add(broken_url)
    logger.info(f"Added new broken URL #{new_id}: {broken_url}")
    return row


def run_watchlist_scan():
    checked = 0
    found_new = 0
    watchlist = db.get_all_watchlist()
    for entry in watchlist:
        status = live_checker.check_url_status(entry["url"], shop=SHOP)
        db.update_watchlist_status(entry["url"], status)
        checked += 1
        if status == 404:
            row = add_new_broken_url(entry["url"])
            if row is not None:
                found_new += 1

    result = {"checked": checked, "found_404s": found_new}
    logger.info(f"Watchlist scan completed: Checked {checked} URLs, found {found_new} 404s.")
    return result


def _task_worker_loop():
    """Persistent SQLite Task Queue Worker Thread."""
    logger.info("Background SQLite Task Worker loop started.")
    interval_minutes = max(CFG.get("scan_interval_minutes", 60), 5)
    last_scan_time = 0

    while True:
        try:
            task = db.dequeue_task()
            if task:
                task_id = task["id"]
                task_type = task["task_type"]
                payload = task["payload"]
                logger.info(f"Processing background task #{task_id} ({task_type})")
                try:
                    if task_type == "verify_applied_redirects":
                        time.sleep(payload.get("delay_seconds", 12))
                        item_ids = payload.get("item_ids", [])
                        for i_id in item_ids:
                            m = db.get_match_by_id(i_id)
                            if m:
                                code, label = live_checker.check_live_url_status(m["broken_url"], shop=SHOP)
                                db.update_match_fields(i_id, http_status=code, http_label=label)
                    elif task_type == "run_watchlist_scan":
                        run_watchlist_scan()
                    db.complete_task(task_id)
                except Exception as ex:
                    logger.error(f"Task #{task_id} failed: {ex}")
                    db.complete_task(task_id, error=str(ex))

            now = time.time()
            if now - last_scan_time > (interval_minutes * 60):
                last_scan_time = now
                if db.get_all_watchlist():
                    run_watchlist_scan()

        except Exception as e:
            logger.error(f"Error in task worker loop: {e}")

        time.sleep(2)


def schedule_delayed_verification(item_ids, delay_seconds=12):
    if item_ids:
        db.enqueue_task("verify_applied_redirects", {"item_ids": item_ids, "delay_seconds": delay_seconds})


def extract_urls_from_rows(rows_data):
    urls = []
    if not rows_data:
        return urls

    first_row = [str(cell or "").strip().lower() for cell in rows_data[0]]
    target_col_idx = None
    for idx, header_val in enumerate(first_row):
        if any(k in header_val for k in ["broken", "url", "link", "path", "404", "page", "source"]):
            target_col_idx = idx
            break

    start_row = 1 if target_col_idx is not None else 0
    seen = set()

    for row in rows_data[start_row:]:
        if target_col_idx is not None and target_col_idx < len(row):
            val = str(row[target_col_idx] or "").strip()
            if val and val.lower() not in ["url", "broken url", "broken_url", "link", "404", "path"]:
                if val not in seen:
                    urls.append(val)
                    seen.add(val)
        else:
            for cell in row:
                val = str(cell or "").strip()
                if val and (val.startswith("http://") or val.startswith("https://") or val.startswith("/") or ".com/" in val or ".in/" in val):
                    if val.lower() not in ["url", "broken url", "link", "path"]:
                        if val not in seen:
                            urls.append(val)
                            seen.add(val)
                        break
    return urls


# ---------------------------------------------------------------------------
# Flask Setup & Security Middleware
# ---------------------------------------------------------------------------
app = Flask(__name__, template_folder="templates", static_folder="static")

# Security Configuration
app.secret_key = secrets.token_bytes(32)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB Payload Limit
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


@app.before_request
def authenticate_and_authorize_requests():
    # Grant session token for UI page loads
    if not request.path.startswith("/api/"):
        session["auth_token"] = APP_SECRET
        return

    # Uniform API Authentication across all GET/POST/PUT/DELETE endpoints
    req_secret = request.headers.get("X-App-Secret") or session.get("auth_token")
    if req_secret != APP_SECRET:
        logger.warning(f"Forbidden request to {request.path} from {request.remote_addr} — invalid auth token.")
        return jsonify({"ok": False, "error": "Unauthorized request"}), 403


@app.errorhandler(413)
def payload_too_large(error):
    logger.warning("Request payload exceeded 16 MB max size limit.")
    return jsonify({"ok": False, "error": "Uploaded payload is too large. Maximum file limit is 16 MB."}), 413


@app.errorhandler(validators.ValidationError)
def handle_validation_error(e):
    logger.warning(f"Validation error on {request.path}: {e}")
    return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/favicon.ico")
@app.route("/favicon.svg")
def favicon():
    return send_from_directory("static", "favicon.svg", mimetype="image/svg+xml")


@app.route("/")
def index():
    return render_template("review.html", app_secret=APP_SECRET, active_page="review")


@app.route("/dashboard")
def dashboard():
    return render_template("dashboard.html", app_secret=APP_SECRET, active_page="dashboard")


@app.route("/automation")
def automation():
    return render_template("automation.html", app_secret=APP_SECRET, active_page="automation")


# ---------------------------------------------------------------------------
# API Routes
# ---------------------------------------------------------------------------
@app.route("/api/state")
def api_state():
    matches = db.get_all_matches()
    return jsonify({
        "matches": matches,
        "shop": SHOP,
        "configured": CONFIGURED,
        "shopify_health": SHOPIFY_API_HEALTH,
    })


@app.route("/api/set_status", methods=["POST"])
def api_set_status():
    body = request.get_json() or {}
    validators.validate_set_status_payload(body)

    m = db.get_match_by_id(body["id"])
    if not m:
        return jsonify({"ok": False, "error": "Match row not found"}), 404

    updates = {}
    if "status" in body:
        updates["status"] = body["status"]
    if "chosen_index" in body:
        updates["chosen_index"] = body["chosen_index"]
    if "custom_target" in body:
        val = str(body["custom_target"] or "").strip()
        updates["custom_target"] = val if val else None

    db.update_match_fields(body["id"], **updates)
    logger.info(f"Updated match #{body['id']} fields: {updates}")
    return jsonify({"ok": True})


@app.route("/api/apply", methods=["POST"])
def api_apply():
    if not CONFIGURED:
        return jsonify({"error": "Store not configured yet — fill in config.json or .env"}), 400

    body = request.get_json() or {}
    validators.validate_apply_payload(body)

    ids = set(body["ids"])
    results = []
    applied_ids = []
    all_matches = db.get_all_matches()

    for m in all_matches:
        if m["id"] in ids:
            target_url = m.get("custom_target")
            if not target_url and m["matches"]:
                chosen = m["matches"][m.get("chosen_index", 0)]
                target_url = chosen["url"]

            if target_url:
                ok, rid, err, status_code = create_or_update_redirect(m["broken_url"], target_url)
                updates = {"status": status_code}
                if ok:
                    updates["applied_redirect_id"] = rid
                    updates["error"] = None
                    updates["skipped_msg"] = None
                    updates["http_status"] = 301
                    updates["http_label"] = "Redirect (301)"
                    applied_ids.append(m["id"])
                elif status_code == "skipped":
                    updates["skipped_msg"] = str(err)
                    updates["applied_redirect_id"] = rid
                else:
                    updates["error"] = str(err)

                db.update_match_fields(m["id"], **updates)
                results.append({"id": m["id"], "ok": ok, "status": status_code, "error": None if (ok or status_code == "skipped") else str(err)})
                time.sleep(0.5)

    schedule_delayed_verification(applied_ids, delay_seconds=12)
    return jsonify({"results": results})


@app.route("/api/upload_broken_links", methods=["POST"])
def api_upload_broken_links():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename or ""
    if not filename:
        return jsonify({"ok": False, "error": "Empty filename"}), 400

    ext = os.path.splitext(filename)[1].lower()
    raw_bytes = file.read()

    rows_data = []
    if ext in (".csv", ".txt"):
        try:
            content = raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            content = raw_bytes.decode("latin-1", errors="ignore")
        reader = csv.reader(io.StringIO(content))
        rows_data = [row for row in reader if row]
    elif ext in (".xlsx", ".xls"):
        if not HAS_OPENPYXL:
            return jsonify({"ok": False, "error": "openpyxl library not installed on server"}), 400
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    if row and any(row):
                        rows_data.append([str(c) if c is not None else "" for c in row])
        except Exception as e:
            return jsonify({"ok": False, "error": f"Failed to parse Excel file: {e}"}), 400
    else:
        return jsonify({"ok": False, "error": "Unsupported file format. Please upload .csv, .xlsx, .xls, or .txt"}), 400

    extracted_urls = extract_urls_from_rows(rows_data)
    if not extracted_urls:
        return jsonify({"ok": False, "error": "No URLs or broken links found in uploaded file"}), 400

    new_added = 0
    all_matches = db.get_all_matches()
    existing_urls_set = {m["broken_url"] for m in all_matches}
    prods = get_products()
    pats = get_patterns()
    inv_idx = matcher.build_product_index(prods)

    for url in extracted_urls:
        row = add_new_broken_url(url, products=prods, patterns=pats, inverted_index=inv_idx, existing_urls_set=existing_urls_set)
        if row is not None:
            new_added += 1

    logger.info(f"Processed file upload '{filename}': Extracted {len(extracted_urls)} URLs, added {new_added} new matches.")
    return jsonify({
        "ok": True,
        "filename": filename,
        "total_extracted": len(extracted_urls),
        "new_added": new_added,
        "matches_count": len(db.get_all_matches()),
    })


@app.route("/api/instant", methods=["POST"])
def api_instant():
    body = request.get_json() or {}
    validators.validate_instant_payload(body)
    broken_url = body["broken_url"].strip()

    candidate = matcher.extract_candidate_name(broken_url)
    matches = matcher.resolve_matches_for_url(broken_url, get_products(), get_patterns())
    new_id = db.get_next_match_id()

    row = {
        "id": new_id,
        "broken_url": broken_url,
        "extracted_name": candidate,
        "matches": matches,
        "status": "pending",
        "chosen_index": 0,
        "source": "instant_paste",
    }
    db.upsert_match(row)
    logger.info(f"Instant fix added broken link #{new_id}: {broken_url}")
    return jsonify(row)


@app.route("/api/apply_one", methods=["POST"])
def api_apply_one():
    if not CONFIGURED:
        return jsonify({"ok": False, "error": "Store not configured yet — fill in config.json or .env"}), 400

    body = request.get_json() or {}
    row_id = body.get("id")
    if not isinstance(row_id, int):
        return jsonify({"ok": False, "error": "Field 'id' must be integer"}), 400

    m = db.get_match_by_id(row_id)
    if not m:
        return jsonify({"ok": False, "error": "row not found"}), 404

    target_url = m.get("custom_target")
    if not target_url and m["matches"]:
        chosen = m["matches"][m.get("chosen_index", 0)]
        target_url = chosen["url"]

    if not target_url:
        return jsonify({"ok": False, "error": "No match or custom target set"}), 400

    ok, rid, err, status_code = create_or_update_redirect(m["broken_url"], target_url)
    updates = {"status": status_code}
    if ok:
        updates["applied_redirect_id"] = rid
        updates["error"] = None
        updates["skipped_msg"] = None
        updates["http_status"] = 301
        updates["http_label"] = "Redirect (301)"
        schedule_delayed_verification([m["id"]], delay_seconds=12)
    elif status_code == "skipped":
        updates["skipped_msg"] = str(err)
        updates["applied_redirect_id"] = rid
    else:
        updates["error"] = str(err)

    db.update_match_fields(m["id"], **updates)
    return jsonify({"ok": ok, "status": status_code, "error": None if (ok or status_code == "skipped") else str(err), "skipped_msg": updates.get("skipped_msg")})


@app.route("/api/check_statuses", methods=["POST"])
def api_check_statuses():
    body = request.get_json() or {}
    validators.validate_check_statuses_payload(body)
    target_ids = set(body.get("ids", []))
    all_matches = db.get_all_matches()

    items_to_check = [
        m for m in all_matches
        if not target_ids or m["id"] in target_ids
    ][:100]  # Hard cap max 100 items per request

    def _check_one(item):
        time.sleep(0.12)  # Micro-pacing delay to prevent storefront anti-bot rate limits
        code, label = live_checker.check_live_url_status(item["broken_url"], shop=SHOP)
        return item["id"], code, label

    with ThreadPoolExecutor(max_workers=4) as executor:
        results = list(executor.map(_check_one, items_to_check))

    for item_id, code, label in results:
        db.update_match_fields(item_id, http_status=code, http_label=label)

    logger.info(f"Checked HTTP statuses for {len(results)} URLs.")
    return jsonify({"ok": True, "checked": len(results), "total": len(all_matches)})


@app.route("/api/rematch", methods=["POST"])
def api_rematch():
    body = request.get_json() or {}
    refetch_catalog = body.get("refetch_catalog", False)
    sync_catalog_only = body.get("sync_catalog_only", False)
    offset = int(body.get("offset", 0))
    limit = int(body.get("limit", 0))

    products_refetched = False

    if refetch_catalog and (offset == 0 or sync_catalog_only):
        if auth.is_configured(CFG):
            try:
                fresh_products = catalog_fetcher.fetch_all_products(CFG)
                if fresh_products:
                    db.replace_all_products(fresh_products)
                    products_refetched = True
                    logger.info(f"Refetched {len(fresh_products)} active products from Shopify API.")
            except Exception as e:
                logger.error(f"Failed to refetch catalog: {e}")
                return jsonify({"ok": False, "error": f"Failed to refetch catalog: {e}"}), 500

    all_matches = db.get_all_matches()
    target_items = [m for m in all_matches if m.get("status") in ("pending", "rejected", "ignored")]
    total_target = len(target_items)

    if sync_catalog_only:
        return jsonify({
            "ok": True,
            "catalog_synced": True,
            "products_count": len(get_products()),
            "total_pending": total_target,
            "products_refetched": products_refetched,
        })

    if limit > 0:
        target_items = target_items[offset : offset + limit]

    products = get_products()
    patterns = get_patterns()
    inv_index = matcher.build_product_index(products)
    count = 0

    for m in target_items:
        matches = matcher.resolve_matches_for_url(m["broken_url"], products, patterns, inverted_index=inv_index)
        db.update_match_fields(m["id"], matches_json=json.dumps(matches), chosen_index=0)
        count += 1

    logger.info(f"Rematched batch {offset}-{offset + count} of {total_target} broken URLs.")
    return jsonify({
        "ok": True,
        "rematched_count": count,
        "offset": offset,
        "processed": offset + len(target_items),
        "total": total_target,
        "products_count": len(products),
        "products_refetched": products_refetched,
    })


@app.route("/api/kpis")
def api_kpis():
    all_matches = db.get_all_matches()
    total = len(all_matches)
    by_status = {"pending": 0, "confirmed": 0, "applied": 0, "rejected": 0, "error": 0, "skipped": 0, "ignored": 0}
    conf_buckets = {"high": 0, "mid": 0, "low": 0, "none": 0}
    scores = []
    rows = []

    for m in all_matches:
        st = m.get("status", "pending")
        by_status[st] = by_status.get(st, 0) + 1
        top = m["matches"][0] if m.get("matches") else None
        score = top["score"] if top else None
        if score is None:
            conf_buckets["none"] += 1
        elif score >= 0.7:
            conf_buckets["high"] += 1
        elif score >= 0.45:
            conf_buckets["mid"] += 1
        else:
            conf_buckets["low"] += 1
        if score is not None:
            scores.append(score)
        rows.append({
            "id": m["id"],
            "broken_url": m["broken_url"],
            "matched_title": top["title"] if top else (m.get("custom_target") or None),
            "matched_url": top["url"] if top else (m.get("custom_target") or None),
            "score": score,
            "status": st,
        })

    avg_score = round(sum(scores) / len(scores), 3) if scores else 0
    fixed_pct = round(100 * by_status["applied"] / total, 1) if total else 0

    return jsonify({
        "total": total,
        "by_status": by_status,
        "conf_buckets": conf_buckets,
        "avg_score": avg_score,
        "fixed_pct": fixed_pct,
        "rows": rows,
        "shop": SHOP,
        "configured": CONFIGURED,
    })


@app.route("/api/patterns", methods=["GET"])
def api_patterns_get():
    return jsonify(db.get_all_patterns())


@app.route("/api/patterns", methods=["POST"])
def api_patterns_post():
    body = request.get_json() or {}
    validators.validate_pattern_payload(body)
    pat = db.add_pattern(body["from_pattern"].strip(), body["to_template"].strip())
    logger.info(f"Added pattern rule #{pat['id']}: {pat['from_pattern']} -> {pat['to_template']}")
    return jsonify(pat)


@app.route("/api/patterns/<int:pat_id>/toggle", methods=["POST"])
def api_patterns_toggle(pat_id):
    enabled = db.toggle_pattern(pat_id)
    if enabled is None:
        return jsonify({"error": "not found"}), 404
    logger.info(f"Toggled pattern rule #{pat_id} -> enabled={enabled}")
    return jsonify({"id": pat_id, "enabled": enabled})


@app.route("/api/patterns/<int:pat_id>", methods=["DELETE"])
def api_patterns_delete(pat_id):
    db.delete_pattern(pat_id)
    logger.info(f"Deleted pattern rule #{pat_id}")
    return jsonify({"ok": True})


@app.route("/api/watchlist", methods=["GET"])
def api_watchlist_get():
    return jsonify(db.get_all_watchlist())


@app.route("/api/watchlist", methods=["POST"])
def api_watchlist_post():
    body = request.get_json() or {}
    validators.validate_watchlist_payload(body)
    db.add_watchlist_url(body["url"].strip())
    logger.info(f"Added watchlist URL: {body['url'].strip()}")
    return jsonify({"ok": True})


@app.route("/api/watchlist/import_sitemap", methods=["POST"])
def api_watchlist_import_sitemap():
    body = request.get_json() or {}
    validators.validate_sitemap_payload(body)
    sitemap_url = body["sitemap_url"].strip()

    try:
        urls = fetch_sitemap_urls(sitemap_url)
    except Exception as e:
        logger.error(f"Failed to parse sitemap at {sitemap_url}: {e}")
        return jsonify({"error": f"Couldn't fetch sitemap: {e}"}), 400

    existing_urls = {w["url"] for w in db.get_all_watchlist()}
    new_urls = [u for u in urls if u not in existing_urls]
    db.bulk_add_watchlist(new_urls)

    logger.info(f"Imported sitemap {sitemap_url}: Found {len(urls)} URLs, added {len(new_urls)} new watchlist entries.")
    return jsonify({"ok": True, "found": len(urls), "added": len(new_urls), "total": len(db.get_all_watchlist())})


@app.route("/api/watchlist/scan_now", methods=["POST"])
def api_watchlist_scan_now():
    watchlist = db.get_all_watchlist()
    if not watchlist:
        return jsonify({"error": "Watchlist is empty — add URLs or import a sitemap first"}), 400
    result = run_watchlist_scan()
    return jsonify({"ok": True, **result})


def build_export_rows(status_filter="all"):
    rows = []
    all_matches = db.get_all_matches()
    for m in all_matches:
        if status_filter != "all" and m.get("status", "pending") != status_filter:
            continue
        chosen = m["matches"][m.get("chosen_index", 0)] if m.get("matches") else None
        target_url = m.get("custom_target") or (chosen["url"] if chosen else "")
        target_title = ("🎯 Custom: " + m["custom_target"]) if m.get("custom_target") else (chosen["title"] if chosen else "")
        others = [
            f"{om['title']} ({round(om['score']*100)}%)"
            for i, om in enumerate(m.get("matches", [])) if i != m.get("chosen_index", 0)
        ]
        rows.append({
            "broken_url": m["broken_url"],
            "extracted_name": m.get("extracted_name", ""),
            "matched_title": target_title,
            "matched_url": target_url,
            "confidence_pct": round(chosen["score"] * 100, 1) if (chosen and not m.get("custom_target")) else ("Custom" if m.get("custom_target") else ""),
            "status": m.get("status", "pending"),
            "other_candidates": "; ".join(others),
        })
    return rows


EXPORT_HEADERS = [
    ("broken_url", "Broken URL"),
    ("matched_url", "New URL (redirect target)"),
    ("matched_title", "Matched Product"),
    ("confidence_pct", "Confidence %"),
    ("status", "Review Status"),
    ("extracted_name", "Extracted Name (from broken URL)"),
    ("other_candidates", "Other Candidates Considered"),
]


@app.route("/api/export.csv")
def export_csv():
    status_filter = request.args.get("filter", "all")
    rows = build_export_rows(status_filter)

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([label for _, label in EXPORT_HEADERS])
    for r in rows:
        writer.writerow([r[key] for key, _ in EXPORT_HEADERS])

    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=broken_link_redirects_{status_filter}.csv"},
    )


@app.route("/api/export.xlsx")
def export_xlsx():
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl not installed — install it or use .csv export"}), 500

    status_filter = request.args.get("filter", "all")
    rows = build_export_rows(status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Redirects"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="171A21", end_color="171A21", fill_type="solid")
    for col, (_, label) in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    status_fills = {
        "applied": PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid"),
        "confirmed": PatternFill(start_color="E6F9EF", end_color="E6F9EF", fill_type="solid"),
        "rejected": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "ignored": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "error": PatternFill(start_color="FFEAEA", end_color="FFEAEA", fill_type="solid"),
        "pending": None,
    }

    for r_idx, r in enumerate(rows, start=2):
        fill = status_fills.get(r["status"])
        for col, (key, _) in enumerate(EXPORT_HEADERS, start=1):
            cell = ws.cell(row=r_idx, column=col, value=r[key])
            if fill:
                cell.fill = fill

    widths = [45, 45, 30, 12, 14, 30, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"broken_link_redirects_{status_filter}.xlsx",
    )


# ---------------------------------------------------------------------------
# Application Entrypoint & Worker Launcher
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    products = get_products()
    matches = db.get_all_matches()
    patterns = get_patterns()
    watchlist = db.get_all_watchlist()

    logger.info(f"Loaded {len(products)} products, {len(matches)} matches, {len(patterns)} patterns, {len(watchlist)} watchlist URLs from SQLite DB.")
    logger.info(f"Shop Domain: {SHOP or '(not set)'} | Configured: {CONFIGURED}")

    # Launch background SQLite task worker thread
    threading.Thread(target=_task_worker_loop, daemon=True).start()

    logger.info("Open http://localhost:5000 in your browser.")
    app.run(host="127.0.0.1", port=5000, debug=False)